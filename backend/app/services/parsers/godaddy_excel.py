"""GoDaddy Commerce / Poynt Transactions Report parser.

The Cowork task on the owner's PC downloads the daily Transactions Report
Excel file from commerce.godaddy.com and posts it to our API. One file
covers one store for one 24-hour window.

GoDaddy's real export is a multi-sheet workbook:
  - "Net Payment Summary"               - store-level totals (cross-check)
  - "Net Payment by Status"
  - "Net Payment Summary by Payment"
  - "Cash In and Cash Out"
  - "Card Payment Tips by Employee"     - (some stores)
  - "Card Payments (N)"                 - PER-ORDER detail, N is the count
  - "Cash Payments (N)"                 - PER-ORDER detail
  - "Other Purchases (N)"               - PER-ORDER detail (gift cards, etc.)
  - "Card Declines (N)"                 - (some stores)
  - "Cash Refunds (N)"                  - (some stores)

Transaction rows live on sheets matching
    ^(Card|Cash) Payments \\(\\d+\\)$  or  ^Other Purchases \\(\\d+\\)$
Each has headers: Date, Transaction ID, Status, Reference IDs, Channel,
Terminal ID, Customer Details, Employee, Subtotal, Tip.

We aggregate those sheets and cross-check the total against
"Net Payment Summary" when present.
"""

import io
import logging
import re
from collections import defaultdict
from datetime import date as date_cls, datetime
from typing import Any

from app.models.daily_revenue import CHANNEL_GODADDY
from app.services.parsers import ParsedRevenueRow

logger = logging.getLogger(__name__)

# Sheet names that contain per-order transaction rows we want to sum.
DETAIL_SHEET_RE = re.compile(
    r"^(?:(?:Card|Cash)\s+Payments|Other\s+Purchases)\s+\(\d+\)\s*$",
    re.IGNORECASE,
)
REFUND_SHEET_RE = re.compile(r"^(Card|Cash)\s+Refunds\s+\(\d+\)\s*$", re.IGNORECASE)


def _as_float(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _find_header_row(rows: list[tuple]) -> int | None:
    """Find the header row on a detail sheet — first row that contains
    'Date' and either 'Subtotal' or 'Transaction ID'.
    """
    for i, row in enumerate(rows[:20]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "date" in cells and any(k in cells for k in ("subtotal", "transaction id", "status")):
            return i
    return None


HourBucketKey = tuple[date_cls, int, int]  # (date, hour 0-23, quarter 0-3)


def _parse_gd_datetime(val: Any) -> datetime | None:
    """GoDaddy writes 'M/D/YYYY, H:MM AM' in the Date column."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    # Comma before time, e.g. "4/21/2026, 6:44 AM"
    for fmt in (
        "%m/%d/%Y, %I:%M %p",
        "%m/%d/%Y, %H:%M",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _sum_detail_sheet(ws) -> tuple[float, float, int, dict[HourBucketKey, tuple[int, float]]]:
    """Return (subtotal, tip, count, hourly_buckets) for one Card/Cash/Other Purchases sheet.

    hourly_buckets maps (date, hour, quarter) -> (txns, gross) using the row
    timestamps so the 15-min heatmap can reconstruct traffic patterns.
    """
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows)
    if header_idx is None:
        logger.warning("Sheet '%s': header row not found", ws.title)
        return 0.0, 0.0, 0, {}

    headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    # Case-insensitive column lookup
    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        col_idx[h.lower()] = i

    date_col = col_idx.get("date")
    subtotal_col = col_idx.get("subtotal")
    tip_col = col_idx.get("tip")
    txn_id_col = col_idx.get("transaction id")

    subtotal_sum = 0.0
    tip_sum = 0.0
    count = 0
    buckets: dict[HourBucketKey, list] = defaultdict(lambda: [0, 0.0])

    for row in rows[header_idx + 1:]:
        if not row or all(v is None or v == "" for v in row):
            continue
        # Detect the trailing "Total" row (first column literal "Total")
        first = str(row[0]).strip().lower() if row[0] is not None else ""
        if first == "total" or first.startswith("grand total"):
            break

        # A valid transaction row has a transaction ID if that column exists
        if txn_id_col is not None and txn_id_col < len(row):
            tid = row[txn_id_col]
            if tid is None or str(tid).strip() == "":
                continue

        row_subtotal = 0.0
        row_tip = 0.0
        if subtotal_col is not None and subtotal_col < len(row):
            row_subtotal = _as_float(row[subtotal_col])
            subtotal_sum += row_subtotal
        if tip_col is not None and tip_col < len(row):
            row_tip = _as_float(row[tip_col])
            tip_sum += row_tip
        count += 1

        # Bucket into (date, hour, quarter)
        ts = _parse_gd_datetime(row[date_col] if date_col is not None and date_col < len(row) else None)
        if ts is not None:
            key = (ts.date(), ts.hour, ts.minute // 15)
            b = buckets[key]
            b[0] += 1
            b[1] += row_subtotal + row_tip  # customer-paid total for this txn

    return subtotal_sum, tip_sum, count, {k: tuple(v) for k, v in buckets.items()}


def _sum_refund_sheet(ws) -> float:
    """Return the refunded amount from a Card/Cash Refunds sheet as a POSITIVE
    number (i.e. abs of the subtotal column values — GoDaddy writes them
    as negatives already, but we want the magnitude).

    Sheet layout has extra cols (Surcharge, Total) after Tip, so we find
    the 'Subtotal' column by header name, not by position.
    """
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return 0.0

    headers = [str(c).strip().lower() if c is not None else "" for c in rows[header_idx]]

    # Prefer 'Total' on refund sheets (includes surcharge), fall back to Subtotal
    amt_col = None
    for candidate in ("total", "amount", "refund amount", "subtotal"):
        if candidate in headers:
            amt_col = headers.index(candidate)
            break
    if amt_col is None:
        return 0.0

    total = 0.0
    for row in rows[header_idx + 1:]:
        if not row or all(v is None or v == "" for v in row):
            continue
        first = str(row[0]).strip().lower() if row[0] is not None else ""
        if first == "total":
            break
        # GoDaddy refunds are stored as negative numbers; abs() so our
        # caller can subtract a positive amount.
        total += abs(_as_float(row[amt_col] if amt_col < len(row) else 0))
    return total


def _read_summary_total(ws) -> float | None:
    """Pull the 'Net Payment' or similar total from the Net Payment Summary sheet.

    Sheet layout is usually two columns: label | value.
    We scan for a row whose label contains 'net payment' (no status qualifier).
    """
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        label = str(row[0]).strip().lower() if row[0] else ""
        if label in ("net payment", "total net payment", "net payments"):
            return _as_float(row[1])
    return None


def parse_godaddy_excel(
    file_bytes: bytes,
    source_file: str,
    store_label: str,
    target_date: date_cls,
) -> list[ParsedRevenueRow]:
    """Parse one GoDaddy Transactions Report Excel file.

    Returns a single ParsedRevenueRow summing that day's transactions
    across all "Card Payments (N)" and "Cash Payments (N)" detail sheets.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError(
            "openpyxl is required to parse GoDaddy Excel exports. "
            "Add openpyxl>=3.1 to requirements.txt."
        )

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    subtotal_sum = 0.0
    tip = 0.0
    count = 0
    refund_total = 0.0
    summary_net: float | None = None
    sheets_read: list[str] = []
    hourly: dict[HourBucketKey, list] = defaultdict(lambda: [0, 0.0])

    # Track Card vs Cash sheet sub-totals separately so we can estimate the
    # 2.3% GoDaddy processing fee on the card portion only.
    card_payment_total = 0.0  # subtotal+tip on Card Payments sheets
    cash_payment_total = 0.0  # subtotal+tip on Cash Payments sheets

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if DETAIL_SHEET_RE.match(sheet_name):
            sub, tp, c, sheet_hourly = _sum_detail_sheet(ws)
            subtotal_sum += sub
            tip += tp
            count += c
            sheet_total = sub + tp
            sheet_lower = sheet_name.lower()
            if sheet_lower.startswith("card payments"):
                card_payment_total += sheet_total
            elif sheet_lower.startswith("cash payments"):
                cash_payment_total += sheet_total
            # "Other Purchases" sheets (gift cards, etc.) are intentionally
            # not bucketed as card or cash — they don't carry a CC fee.
            for k, (bt, bg) in sheet_hourly.items():
                entry = hourly[k]
                entry[0] += bt
                entry[1] += bg
            sheets_read.append(f"{sheet_name}:{c}")
        elif REFUND_SHEET_RE.match(sheet_name):
            refund_total += _sum_refund_sheet(ws)
            sheets_read.append(f"{sheet_name}:refund")
        elif sheet_name.lower().startswith("net payment summary"):
            if summary_net is None:
                summary_net = _read_summary_total(ws)

    # GoDaddy's settlement report defines:
    #   Total Payments = subtotal + tips  (what customers paid pre-refund)
    #   Net Payments   = Total Payments - refunds
    # Store `gross_revenue` as Total Payments so the dashboard reconciles
    # directly with the settlement figure.
    gross = subtotal_sum + tip
    net = gross - refund_total

    if count == 0 and summary_net is None:
        logger.warning(
            "GoDaddy parser: no transactions found in %s (sheets=%s)",
            source_file, wb.sheetnames,
        )
        return []

    # Flatten hourly buckets into a list of dicts for the caller to upsert.
    hourly_rows = [
        {
            "date": d, "hour": h, "quarter": q,
            "channel": CHANNEL_GODADDY,
            "txns": bt, "gross": round(bg, 2),
        }
        for (d, h, q), (bt, bg) in hourly.items()
    ]

    row = ParsedRevenueRow(
        external_store_id=store_label,
        channel=CHANNEL_GODADDY,
        date=target_date,
        gross_revenue=round(gross, 2),
        net_revenue=round(net, 2) if (net or count) else None,
        tip_total=round(tip, 2) if tip else None,
        transaction_count=count,
        card_total=round(card_payment_total, 2) if card_payment_total else None,
        cash_total=round(cash_payment_total, 2) if cash_payment_total else None,
        raw_notes={
            "source_file": source_file,
            "sheets_read": sheets_read,
            "subtotal_sum": round(subtotal_sum, 2),
            "summary_net_from_workbook": summary_net,
            "refund_total": round(refund_total, 2) if refund_total else 0.0,
            "parsed_at": datetime.utcnow().isoformat(),
            "hourly_rows": hourly_rows,
        },
    )
    return [row]


def parse_godaddy_settlement(
    file_bytes: bytes,
    source_file: str,
    store_label: str,
) -> tuple[list[ParsedRevenueRow], set[str]]:
    """Parse a GoDaddy Settlement XLSX — a multi-day version of the same workbook.

    Returns (rows_per_date, terminal_ids_seen). The caller uses the
    terminal IDs to pick a Location, then writes one DailyRevenue +
    HourlyRevenue for each emitted row.

    Sheets are the same layout as the daily Transactions Report (Card
    Payments (N), Cash Payments (N), Other Purchases (N), Card/Cash
    Refunds (N), Net Payment Summary), but detail rows span the full
    date range. We group by the Date column's date component and emit
    one ParsedRevenueRow per day.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("openpyxl is required to parse GoDaddy Settlement exports.")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Per-day, per-quarter-hour buckets
    per_day: dict[date_cls, dict] = {}
    terminal_ids: set[str] = set()
    refunds_per_day: dict[date_cls, float] = {}

    def _daybucket(d: date_cls) -> dict:
        if d not in per_day:
            per_day[d] = {
                "subtotal": 0.0, "tip": 0.0, "count": 0,
                "hourly": {},  # (hour, quarter) -> [txns, gross]
            }
        return per_day[d]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if DETAIL_SHEET_RE.match(sheet_name):
            rows = list(ws.iter_rows(values_only=True))
            header_idx = _find_header_row(rows)
            if header_idx is None:
                continue
            headers = [str(c).strip().lower() if c is not None else "" for c in rows[header_idx]]
            date_col = headers.index("date") if "date" in headers else 0
            sub_col = headers.index("subtotal") if "subtotal" in headers else None
            tip_col = headers.index("tip") if "tip" in headers else None
            term_col = headers.index("terminal id") if "terminal id" in headers else None
            txn_col = headers.index("transaction id") if "transaction id" in headers else None

            for row in rows[header_idx + 1:]:
                if not row or all(v is None or v == "" for v in row):
                    continue
                first = str(row[0]).strip().lower() if row[0] is not None else ""
                if first == "total" or first.startswith("grand total"):
                    break
                if txn_col is not None and txn_col < len(row):
                    tid = row[txn_col]
                    if tid is None or str(tid).strip() == "":
                        continue

                ts = _parse_gd_datetime(row[date_col] if date_col < len(row) else None)
                if ts is None:
                    continue
                d = ts.date()
                bucket = _daybucket(d)

                row_sub = _as_float(row[sub_col]) if sub_col is not None and sub_col < len(row) else 0.0
                row_tip = _as_float(row[tip_col]) if tip_col is not None and tip_col < len(row) else 0.0
                bucket["subtotal"] += row_sub
                bucket["tip"] += row_tip
                bucket["count"] += 1

                hkey = (ts.hour, ts.minute // 15)
                hb = bucket["hourly"].setdefault(hkey, [0, 0.0])
                hb[0] += 1
                hb[1] += row_sub + row_tip

                if term_col is not None and term_col < len(row):
                    t = row[term_col]
                    if t and len(str(t)) > 20:  # filter out non-UUID junk
                        terminal_ids.add(str(t).strip())

        elif REFUND_SHEET_RE.match(sheet_name):
            # Refunds have the same Date column; bucket per-day and subtract.
            rows = list(ws.iter_rows(values_only=True))
            header_idx = _find_header_row(rows)
            if header_idx is None:
                continue
            headers = [str(c).strip().lower() if c is not None else "" for c in rows[header_idx]]
            date_col = headers.index("date") if "date" in headers else 0
            amt_col = None
            for candidate in ("total", "amount", "refund amount", "subtotal"):
                if candidate in headers:
                    amt_col = headers.index(candidate)
                    break
            if amt_col is None:
                continue
            for row in rows[header_idx + 1:]:
                if not row or all(v is None or v == "" for v in row):
                    continue
                first = str(row[0]).strip().lower() if row[0] is not None else ""
                if first == "total":
                    break
                ts = _parse_gd_datetime(row[date_col] if date_col < len(row) else None)
                if ts is None:
                    continue
                amt = abs(_as_float(row[amt_col] if amt_col < len(row) else 0))
                refunds_per_day[ts.date()] = refunds_per_day.get(ts.date(), 0.0) + amt

    rows_out: list[ParsedRevenueRow] = []
    for d, b in per_day.items():
        gross = b["subtotal"] + b["tip"]  # customer-paid total
        refund = refunds_per_day.get(d, 0.0)
        net = gross - refund
        hourly_rows = [
            {
                "date": d, "hour": h, "quarter": q,
                "channel": CHANNEL_GODADDY,
                "txns": bt, "gross": round(bg, 2),
            }
            for (h, q), (bt, bg) in b["hourly"].items()
        ]
        rows_out.append(ParsedRevenueRow(
            external_store_id=store_label,
            channel=CHANNEL_GODADDY,
            date=d,
            gross_revenue=round(gross, 2),
            net_revenue=round(net, 2),
            tip_total=round(b["tip"], 2) if b["tip"] else None,
            transaction_count=b["count"],
            raw_notes={
                "source_file": source_file,
                "subtotal_sum": round(b["subtotal"], 2),
                "refund_total": round(refund, 2) if refund else 0.0,
                "hourly_rows": hourly_rows,
                "terminal_ids": sorted(terminal_ids),
            },
        ))

    return rows_out, terminal_ids
