"""Parser for the GoDaddy Items Sales report XLSX.

Source: commerce.godaddy.com → Reports → Items → Export to XLSX.

Workbook layout:
  - "Summary"      - aggregated by item (we ignore; we re-derive from rows)
  - "Item Details" - one row per item line on each transaction. THIS is
                     what we ingest.

Item Details columns:
  Date | Transaction ID | Order ID | Name/SKU | Unit Price | Quantity |
  Subtotal | Item Discount | Item Fee | Total Taxes | Grand Total | Status

Name/SKU is packed:
  ITEM_NAME, • GROUP: VALUE, • GROUP: VALUE, …, SKU

The bullet character is U+2022. The last comma-separated segment that
DOES NOT start with the bullet is the SKU. Earlier bullet segments are
modifiers in `GROUP: VALUE` form.
"""

from __future__ import annotations

import hashlib
import io
import logging
from dataclasses import dataclass, field
from datetime import datetime

logger = logging.getLogger(__name__)

BULLET = "•"


@dataclass
class ParsedModifier:
    group_name: str
    value: str


@dataclass
class ParsedSaleLine:
    sale_datetime: datetime
    transaction_id: str
    order_id: str | None
    sku: str | None
    item_name: str
    raw_modifier_text: str
    unit_price: float
    quantity: float
    subtotal: float
    item_discount: float
    item_fee: float
    total_taxes: float
    grand_total: float
    status: str | None
    modifiers: list[ParsedModifier] = field(default_factory=list)
    dedup_hash: str = ""


def _parse_dt(val) -> datetime | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in (
        "%m/%d/%Y, %I:%M %p",
        "%m/%d/%Y, %H:%M",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _to_float(val) -> float:
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _split_name_sku_field(s: str) -> tuple[str, list[ParsedModifier], str | None]:
    """Split the packed Name/SKU column into (item_name, modifiers, sku).

    Strategy: split by ',' (single comma — the field uses ', •' between
    entries). The first part is item name. Subsequent parts that begin
    with the bullet are modifiers. The trailing part WITHOUT a bullet is
    the SKU (or omitted, in which case we return None).
    """
    if not s:
        return "", [], None
    parts = [p.strip() for p in s.split(",")]
    if not parts:
        return "", [], None

    item_name = parts[0]
    sku: str | None = None

    # Walk from end backward — the last non-bullet, non-empty segment is SKU.
    end_idx = len(parts)
    if len(parts) >= 2 and not parts[-1].startswith(BULLET):
        sku = parts[-1] or None
        end_idx = len(parts) - 1

    modifiers: list[ParsedModifier] = []
    for p in parts[1:end_idx]:
        if not p.startswith(BULLET):
            continue
        body = p[1:].strip()
        if ":" not in body:
            continue
        group, value = body.split(":", 1)
        modifiers.append(ParsedModifier(group_name=group.strip(), value=value.strip()))

    return item_name, modifiers, sku


def _make_hash(line: ParsedSaleLine) -> str:
    payload = (
        f"{line.transaction_id}|{line.sku or ''}|"
        f"{line.sale_datetime.isoformat()}|"
        f"{line.unit_price:.4f}|{line.quantity}|{line.raw_modifier_text}"
    )
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def parse_godaddy_items_xlsx(file_bytes: bytes) -> list[ParsedSaleLine]:
    """Parse an Items report XLSX into a list of sale lines."""
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to parse GoDaddy Items XLSX") from exc

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Item Details" not in wb.sheetnames:
        raise ValueError("Workbook is missing the 'Item Details' sheet — is this the right export?")

    ws = wb["Item Details"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    # Column lookup, case-insensitive
    col = {str(h).strip().lower() if h is not None else "": i for i, h in enumerate(headers)}
    required = {"date", "transaction id", "order id", "name/sku", "unit price",
                "quantity", "subtotal", "item discount", "item fee", "total taxes",
                "grand total", "status"}
    missing = required - set(col.keys())
    if missing:
        logger.warning("Items XLSX missing columns: %s — continuing best-effort", missing)

    lines: list[ParsedSaleLine] = []
    for row in rows:
        if not row or all(v in (None, "") for v in row):
            continue
        dt = _parse_dt(row[col.get("date", 0)])
        if dt is None:
            continue
        tx_id = str(row[col.get("transaction id", 1)] or "").strip()
        if not tx_id:
            continue

        name_sku = row[col.get("name/sku", 3)] or ""
        item_name, modifiers, sku = _split_name_sku_field(str(name_sku))
        # Special case: "Custom Item" rows have no SKU and only a name.
        if not item_name:
            item_name = str(name_sku).strip() or "(unknown)"

        line = ParsedSaleLine(
            sale_datetime=dt,
            transaction_id=tx_id,
            order_id=(str(row[col.get("order id", 2)] or "").strip() or None) if "order id" in col else None,
            sku=sku,
            item_name=item_name,
            raw_modifier_text=str(name_sku),
            unit_price=_to_float(row[col.get("unit price", 4)]),
            quantity=_to_float(row[col.get("quantity", 5)]) or 1.0,
            subtotal=_to_float(row[col.get("subtotal", 6)]),
            item_discount=_to_float(row[col.get("item discount", 7)]),
            item_fee=_to_float(row[col.get("item fee", 8)]),
            total_taxes=_to_float(row[col.get("total taxes", 9)]),
            grand_total=_to_float(row[col.get("grand total", 10)]),
            status=(str(row[col.get("status", 11)] or "").strip() or None) if "status" in col else None,
            modifiers=modifiers,
        )
        line.dedup_hash = _make_hash(line)
        lines.append(line)

    return lines
