"""Monthly expenses — CRUD + one-shot P&L Excel seed.

Routes:
  GET  /expenses                    - all expenses, grouped by location_id
  POST /expenses                    - create
  PATCH /expenses/{id}              - edit
  DELETE /expenses/{id}             - delete
  GET  /expenses/settings           - labor_burden_multiplier + cogs_percent
  PATCH /expenses/settings          - update those
  POST /expenses/seed-from-pnl      - upload Store P&L.xlsx, populate rows
"""

from __future__ import annotations

import io
import logging
import re
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import require_roles
from app.models.expense import Expense
from app.models.location import Location
from app.models.system_settings import SystemSettings
from app.models.user import User, UserRole

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/expenses", tags=["expenses"])

# Only this owner account may mutate expenses or P&L settings. All other
# users — including the Jess owner account — get read-only access. The
# expense table feeds directly into the Elite scorecard profit math, so
# an accidental edit can silently corrupt every shop's margin.
EXPENSES_WRITE_EMAIL = "logcastles@gmail.com"


def _require_expense_writer(current_user: User = Depends(require_roles(UserRole.owner))) -> User:
    if (current_user.email or "").strip().lower() != EXPENSES_WRITE_EMAIL:
        raise HTTPException(
            status_code=403,
            detail="Expenses are locked — only the primary owner can edit them.",
        )
    return current_user

# Categories that come from the P&L but should NEVER be stored as Expense
# rows — they are derived from other ingestion sources instead.
SKIP_CATEGORIES = {
    "payroll",  # derived from Homebase × labor_burden_multiplier
}


# ------------------------- Pydantic schemas ---------------------------

class ExpenseIn(BaseModel):
    location_id: int | None = None
    category: str
    amount: float
    notes: str | None = None


class ExpensePatch(BaseModel):
    category: str | None = None
    amount: float | None = None
    notes: str | None = None


class SettingsPatch(BaseModel):
    labor_burden_multiplier: float | None = None
    cogs_percent: float | None = None


# ------------------------- Helpers ------------------------------------

async def _settings_row(db: AsyncSession) -> SystemSettings:
    row = (await db.execute(select(SystemSettings).limit(1))).scalar_one_or_none()
    if row is None:
        row = SystemSettings(id=1)
        db.add(row)
        await db.flush()
    return row


def _serialize_expense(e: Expense) -> dict:
    return {
        "id": e.id,
        "location_id": e.location_id,
        "category": e.category,
        "amount": e.amount,
        "notes": e.notes,
        "updated_at": e.updated_at.isoformat() + "Z" if e.updated_at else None,
    }


# ------------------------- CRUD ---------------------------------------

@router.get("")
async def list_expenses(
    current_user: User = Depends(require_roles(UserRole.owner)),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(
        select(Expense).order_by(Expense.location_id, Expense.category)
    )).scalars().all()
    locations = (await db.execute(
        select(Location).where(Location.canonical_short_name.isnot(None))
        .order_by(Location.canonical_short_name)
    )).scalars().all()

    can_edit = (current_user.email or "").strip().lower() == EXPENSES_WRITE_EMAIL
    return {
        "expenses": [_serialize_expense(e) for e in rows],
        "locations": [
            {
                "id": loc.id,
                "name": loc.name,
                "canonical_short_name": loc.canonical_short_name,
            }
            for loc in locations
        ],
        "can_edit": can_edit,
    }


@router.post("")
async def create_expense(
    body: ExpenseIn,
    current_user: User = Depends(_require_expense_writer),
    db: AsyncSession = Depends(get_db),
):
    if body.category.strip().lower() in SKIP_CATEGORIES:
        raise HTTPException(
            status_code=400,
            detail="Payroll is derived from Homebase × labor_burden_multiplier; do not add it as an expense row.",
        )
    e = Expense(
        location_id=body.location_id,
        category=body.category.strip(),
        amount=body.amount,
        notes=body.notes,
    )
    db.add(e)
    await db.commit()
    await db.refresh(e)
    return _serialize_expense(e)


# ------------------------- Settings -----------------------------------
# Declared BEFORE the /{expense_id} routes so FastAPI's route matcher
# reaches these first — otherwise PATCH /expenses/settings gets caught
# by the numeric {expense_id} path and returns 422.

@router.get("/settings")
async def get_settings(
    current_user: User = Depends(require_roles(UserRole.owner)),
    db: AsyncSession = Depends(get_db),
):
    row = await _settings_row(db)
    return {
        "labor_burden_multiplier": row.labor_burden_multiplier,
        "cogs_percent": row.cogs_percent,
    }


@router.patch("/settings")
async def update_settings(
    body: SettingsPatch,
    current_user: User = Depends(_require_expense_writer),
    db: AsyncSession = Depends(get_db),
):
    row = await _settings_row(db)
    if body.labor_burden_multiplier is not None:
        if body.labor_burden_multiplier < 1.0 or body.labor_burden_multiplier > 2.0:
            raise HTTPException(status_code=400, detail="labor_burden_multiplier out of range (1.0 - 2.0)")
        row.labor_burden_multiplier = body.labor_burden_multiplier
    if body.cogs_percent is not None:
        if body.cogs_percent < 0 or body.cogs_percent > 1:
            raise HTTPException(status_code=400, detail="cogs_percent must be between 0 and 1")
        row.cogs_percent = body.cogs_percent
    row.updated_at = datetime.utcnow()
    await db.commit()
    return {
        "labor_burden_multiplier": row.labor_burden_multiplier,
        "cogs_percent": row.cogs_percent,
    }


# ------------------------- Single-row CRUD ----------------------------

@router.patch("/{expense_id}")
async def update_expense(
    expense_id: int,
    body: ExpensePatch,
    current_user: User = Depends(_require_expense_writer),
    db: AsyncSession = Depends(get_db),
):
    e = (await db.execute(
        select(Expense).where(Expense.id == expense_id)
    )).scalar_one_or_none()
    if not e:
        raise HTTPException(status_code=404, detail="Expense not found")

    if body.category is not None:
        if body.category.strip().lower() in SKIP_CATEGORIES:
            raise HTTPException(status_code=400, detail="Payroll is derived, not stored here")
        e.category = body.category.strip()
    if body.amount is not None:
        e.amount = body.amount
    if body.notes is not None:
        e.notes = body.notes
    e.updated_at = datetime.utcnow()
    await db.commit()
    return _serialize_expense(e)


@router.delete("/{expense_id}")
async def delete_expense(
    expense_id: int,
    current_user: User = Depends(_require_expense_writer),
    db: AsyncSession = Depends(get_db),
):
    e = (await db.execute(
        select(Expense).where(Expense.id == expense_id)
    )).scalar_one_or_none()
    if not e:
        raise HTTPException(status_code=404, detail="Expense not found")
    await db.delete(e)
    await db.commit()
    return {"ok": True}


# ------------------------- P&L Excel seed -----------------------------

# Maps the per-store column block headers in Store P&L.xlsx to the
# canonical_short_name on the Location table.
PNL_STORE_HEADER_TO_SHORT = {
    "apple valley": "APPLE_VALLEY_HS",
    "victorville": "VICTORVILLE",
    "barstow": "BARSTOW",
    "7th street": "SEVENTH_STREET",
    "warehouse": "WAREHOUSE",
    "bakery": "BAKERY",
    "hesperia": "HESPERIA",
    "yucca loma": "YUCCA_LOMA",
}

# Section-header rows in the P&L that mark a transition between per-store
# expenses and the shared-overhead / COGS sections (location_id = NULL).
SHARED_SECTION_MARKERS = ("overall monthly", "all 6 shops cogs")


def _as_float(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


@router.post("/seed-from-pnl")
async def seed_from_pnl(
    file: UploadFile = File(...),
    replace_existing: bool = False,
    current_user: User = Depends(_require_expense_writer),
    db: AsyncSession = Depends(get_db),
):
    """Seed the `expenses` table from the owner's Store P&L Excel file.

    Parses the multi-column layout where each store gets a 3-column block
    (name, amount, extra) and shared overhead / COGS live in a single
    column block below the per-store data.

    `replace_existing=True` deletes all current Expense rows before
    inserting. `replace_existing=False` only inserts rows that don't
    already exist for that (location_id, category) pair.

    Rows whose category is in SKIP_CATEGORIES (Payroll) are ignored —
    payroll is derived from Homebase × labor_burden_multiplier.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {exc}")

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Resolve location ids by canonical_short_name.
    loc_rows = (await db.execute(select(Location))).scalars().all()
    short_to_id = {loc.canonical_short_name: loc.id for loc in loc_rows if loc.canonical_short_name}

    # Find the "Apple Valley | Victorville | Barstow | ..." header row and
    # build { column_index -> canonical_short_name }.
    store_header_row_idx = None
    col_to_short: dict[int, str] = {}
    for i, row in enumerate(rows[:10]):
        if not row:
            continue
        hits = 0
        tmp: dict[int, str] = {}
        for col, cell in enumerate(row):
            if cell is None:
                continue
            lowered = str(cell).strip().lower()
            if lowered in PNL_STORE_HEADER_TO_SHORT:
                tmp[col] = PNL_STORE_HEADER_TO_SHORT[lowered]
                hits += 1
        if hits >= 3:  # found the real header row
            store_header_row_idx = i
            col_to_short = tmp
            break

    if not col_to_short:
        raise HTTPException(
            status_code=400,
            detail="Could not find the store-name header row in the P&L sheet",
        )

    # Walk rows after the header, grabbing (category, amount) pairs from
    # each per-store column block until we hit the shared-overhead markers.
    to_create: list[tuple[int | None, str, float]] = []
    in_shared_section = False
    in_cogs_section = False  # items inside "All 6 Shops COGS" — SKIPPED, because
                             # cost of goods is modeled as revenue × cogs_percent,
                             # not tracked per-line here.
    shared_col = 0  # Overall Monthly / COGS live in column A

    for i in range(store_header_row_idx + 1, len(rows)):
        row = rows[i]
        if not row:
            continue

        first_val = row[0]
        first_text = (str(first_val).strip() if first_val is not None else "")

        # Section transitions
        lowered_first = first_text.lower()
        if "all 6 shops cogs" in lowered_first:
            in_shared_section = True
            in_cogs_section = True
            continue
        if "overall monthly" in lowered_first:
            in_shared_section = True
            in_cogs_section = False
            continue

        if in_shared_section:
            if in_cogs_section:
                # COGS section — skip. Handled by SystemSettings.cogs_percent.
                continue
            # Shared overhead / COGS — single column block: (category, amount, ...)
            amt = _as_float(row[1]) if len(row) > 1 else None
            if first_text and amt is not None and first_text.lower() not in SKIP_CATEGORIES:
                # Skip informational / computed rows — these aren't real expenses.
                bad = (
                    "total", "monthly net", "projected yearly",
                    "all 6 shops income", "income=",
                    "percentage", "percent",
                    "new costco",  # annotation row
                )
                if not any(k in lowered_first for k in bad):
                    to_create.append((None, first_text, amt))
            continue

        # Per-store blocks — for each mapped column, look at (col, col+1)
        for col, short_name in col_to_short.items():
            if col >= len(row) or col + 1 >= len(row):
                continue
            cat_val = row[col]
            amt_val = row[col + 1] if col + 1 < len(row) else None
            cat_text = (str(cat_val).strip() if cat_val is not None else "")
            if not cat_text:
                continue
            lowered_cat = cat_text.lower()
            if lowered_cat in ("expenses", "income") or lowered_cat.startswith("totals"):
                continue
            if lowered_cat in SKIP_CATEGORIES:
                continue
            amt = _as_float(amt_val)
            if amt is None:
                continue
            loc_id = short_to_id.get(short_name)
            if loc_id is None:
                continue
            to_create.append((loc_id, cat_text, amt))

    if replace_existing:
        # Wipe all current expenses before inserting fresh rows.
        for e in (await db.execute(select(Expense))).scalars().all():
            await db.delete(e)
        await db.flush()

    # Insert, skipping duplicates on (location_id, category) when not replacing.
    existing_keys: set[tuple[int | None, str]] = set()
    if not replace_existing:
        for e in (await db.execute(select(Expense))).scalars().all():
            existing_keys.add((e.location_id, e.category.lower()))

    created = 0
    skipped_dupe = 0
    for loc_id, category, amount in to_create:
        key = (loc_id, category.lower())
        if key in existing_keys:
            skipped_dupe += 1
            continue
        db.add(Expense(location_id=loc_id, category=category, amount=amount))
        existing_keys.add(key)
        created += 1

    await db.commit()
    return {
        "ok": True,
        "parsed": len(to_create),
        "created": created,
        "skipped_duplicates": skipped_dupe,
        "replace_existing": replace_existing,
    }
